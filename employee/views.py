from django.core.exceptions import ValidationError
from django.shortcuts import render, get_object_or_404 ,redirect
from decimal import Decimal
from .models import Quotation, Room, Booking,Enquiry
from django.http import JsonResponse
from django.contrib import messages
from datetime import datetime, timedelta ,date
from weasyprint import HTML
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from django.utils.timezone import now
from django.core.paginator import Paginator
from django.db.models import Sum
import openpyxl
from django.db.models import Count, Sum, Q
import openpyxl
from django.http import HttpResponse
from .models import Booking
from datetime import datetime, timedelta
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from .models import Room, Booking

# Create your views here.


def dashboard(request):
    rooms_with_availability = []

    # Fetch all rooms
    rooms = Room.objects.all()
    total_office=Room.objects.filter(room_type='Office')
    # Fetch only rooms with room_type 'Office'
    office_rooms = Room.objects.filter(room_type='Office')
    total_office_rooms = office_rooms.count()

    total_rooms = total_office.count()
    total_available_sheets = 0
    total_clients_active = Booking.objects.filter(status='ACTIVE', room__room_type='Office').values('client_name').distinct().count()
    total_clients = Booking.objects.values('client_name').distinct().count()
    total_pending_quotations = Quotation.objects.filter(status='Pending').count()

    # Variable to hold total available sheets for office rooms
    available_office_rooms = 0

    for room in rooms:
        active_bookings = Booking.objects.filter(room=room, status='ACTIVE')

        # Calculate booked and available sheets
        booked_sheets = active_bookings.aggregate(Sum('sheets_booked'))['sheets_booked__sum'] or 0
        available_sheets = room.sheet_count - booked_sheets
        total_available_sheets += available_sheets

        # If room is of type 'Office', add available sheets to available_office_rooms
        if room.room_type == 'Office':
            available_office_rooms += available_sheets

        # Append room data to the list
        rooms_with_availability.append({
            'room': room,
            'available_sheets': available_sheets,
            'booked_sheets': booked_sheets,
            'bookings': active_bookings,
        })

    # Prepare context for the template
    context = {
        'rooms_with_availability': rooms_with_availability,
        'total_rooms': total_rooms,
        'total_available_sheets': total_available_sheets,
        'total_clients_active': total_clients_active,
        'total_clients': total_clients,
        'total_pending_quotations': total_pending_quotations,
        'available_office_rooms': available_office_rooms  # Total available sheets for office rooms
    }

    return render(request, 'dashboard.html', context)
def room(request):
    
    rooms = Room.objects.all()
    
    return render(request, 'room.html', {'rooms':rooms}) 

def create_room(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        room_type = request.POST.get('room_type')
        sheet_count = int(request.POST.get('sheet_count'))
        
        # Use default values if the fields are empty
        price_per_sheet = request.POST.get('price_per_sheet')
        price = request.POST.get('price')

        if room_type == "Office":
            price_per_sheet = float(price_per_sheet) if price_per_sheet else 0.0
            price = None  # Office doesn't use the 'price' field
        else:
            price = float(price) if price else 0.0
            price_per_sheet = None  # Other room types don't use 'price_per_sheet'

        # Create and save the room
        Room.objects.create(
            name=name,
            room_type=room_type,
            sheet_count=sheet_count,
            price_per_sheet=price_per_sheet,
            price=price,
        )

        messages.success(request, f"Room '{name}' has been created successfully!")
        return redirect('employee:room')  # Redirect to the room list page

    return render(request, 'create_room.html')

def update_room(request, id):
    # Fetch the room to be updated
    room = get_object_or_404(Room, id=id)

    if request.method == 'POST':
        # Update fields, handling empty strings or missing data
        room.name = request.POST.get('name', room.name)
        room.room_type = request.POST.get('room_type', room.room_type)
        room.sheet_count = int(request.POST.get('sheet_count', room.sheet_count))

        # Handle price_per_sheet safely
        price_per_sheet = request.POST.get('price_per_sheet', None)
        room.price_per_sheet = Decimal(price_per_sheet) if price_per_sheet else room.price_per_sheet

        # Handle price safely
        price = request.POST.get('price', None)
        room.price = Decimal(price) if price else room.price

        room.amenities = request.POST.get('amenities', room.amenities)

        # Save the updated room
        room.save()

        return redirect('employee:room')  # Redirect after successful update

    return render(request, 'update_room.html', {'room': room})


def delete_room(request, id):
    if request.method == "POST":
        # Fetch and delete the room
        room = get_object_or_404(Room, id=id)
        room.delete()
        return JsonResponse({'success': True, 'message': 'Room deleted successfully!'})
    else:
        return JsonResponse({'success': False, 'message': 'Invalid request method.'}, status=400)


    









def room_availability(request, room_id):
    room = get_object_or_404(Room, id=room_id)
    bookings = Booking.objects.filter(room=room).count()
    available_sheets = room.sheet_count - bookings
    return JsonResponse({'available_sheets': available_sheets})


def book_room(request):
    if request.method == 'POST':
        room_id = request.POST['room_id']
        client_name = request.POST['client_name']
        sheets_requested = int(request.POST['sheets_requested'])
        booking_type = request.POST['booking_type']
        start_date = datetime.strptime(request.POST['start_date'], '%Y-%m-%d').date()

        # Determine the end date based on booking type
        if booking_type == 'DAILY':
            end_date = datetime.strptime(request.POST['end_date'], '%Y-%m-%d').date()
        elif booking_type == 'MONTHLY':
            end_date = start_date + timedelta(days=30)
        elif booking_type == 'YEARLY':
            end_date = start_date + timedelta(days=365)

        # Fetch the room and check availability
        room = get_object_or_404(Room, id=room_id)

         # Calculate available sheets based on active bookings
        active_bookings = Booking.objects.filter(room=room, status='ACTIVE')
        booked_sheets = active_bookings.aggregate(Sum('sheets_booked'))['sheets_booked__sum'] or 0
        available_sheets = room.sheet_count - booked_sheets



        # If room type is not "Office", don't calculate based on sheets, just store the price
        if room.room_type != "Office":
            total_price = room.price  # Direct price without calculation based on sheets
        else:
            # For "Office" type, calculate price based on sheets requested
            
            if sheets_requested > available_sheets:
                messages.error(request, "Not enough sheets available for the selected dates.")
                return render(request, 'create_quotation.html', {
                    'rooms': Room.objects.all(),
                    'selected_room': room,
                    'room_type': room.room_type,
                    'price': room.price or room.price_per_sheet,
                    'error_sheets_requested': True,  # Pass error flag
                    'available_sheets': available_sheets,  # Ensure this reflects the correct value
                    'request_data': request.POST, 
                })

            price_per_sheet = room.price_per_sheet or 5000
            total_price = sheets_requested * price_per_sheet


        # Discount and GST calculation
        discount = int(request.POST.get('discount', 0))
        discount_amount = (total_price * discount) / 100
        final_price = total_price - discount_amount

        # Calculate GST
        igst = (final_price * 9) / 100
        sgst = (final_price * 9) / 100
        total_price_with_gst = final_price + igst + sgst

        # Create the booking
        Booking.objects.create(
            room=room,
            client_name=client_name,
            sheets_booked=sheets_requested if room.room_type == "Office" else 0,  # Set sheets only for "Office"
            start_date=start_date,
            end_date=end_date,
            booking_type=booking_type,
            discount=discount,
            final_price=total_price_with_gst,  # Include GST in the final price
        )

        messages.success(request, f"Room booked successfully from {start_date} to {end_date}!")
        return redirect('employee:dashboard')

    else:
        rooms = Room.objects.all()
        selected_room_id = request.GET.get('room_id')
        selected_room = None
        room_type = None
        price = None
        available_sheets = 0
        
        if selected_room_id:
            selected_room = Room.objects.get(id=selected_room_id)
            room_type = selected_room.room_type
            price = selected_room.price or selected_room.price_per_sheet

            # Calculate available sheets dynamically
            active_bookings = Booking.objects.filter(room=selected_room, status='ACTIVE')
            booked_sheets = active_bookings.aggregate(Sum('sheets_booked'))['sheets_booked__sum'] or 0
            available_sheets = selected_room.sheet_count - booked_sheets
            

        else:
            available_sheets = 0
            
        return render(request, 'book_room.html', {
            'rooms': rooms,
            'selected_room': selected_room,
            'room_type': room_type,
            'price': price,
            'error_sheets_requested': False,  # Default to no error
            'available_sheets': available_sheets,
            
        })


def generate_invoice(request, id):
    # Get the booking details
    booking = get_object_or_404(Booking, id=id)
    room = booking.room
    start_date = booking.start_date
    end_date = booking.end_date
    client_name = booking.client_name
    sheets_booked = booking.sheets_booked
    discount = booking.discount  # Assuming you store discount percentage in the booking model

    # Pricing logic
    if room.room_type == 'Auditorium':
        price_per_sheet = 15000
    elif room.room_type == 'Meeting Room':
        price_per_sheet = 1500
    else:
        price_per_sheet = 5000

    # Calculate total cost and final cost
    total_price = price_per_sheet * sheets_booked
    discount_amount = (total_price * discount) / 100
    final_cost = total_price - discount_amount

    # Calculate GST
    igst = (final_cost * 9) / 100
    sgst = (final_cost * 9) / 100
    total_price_with_gst = final_cost + igst + sgst

    # Create a context to pass to the invoice template
    context = {
        'client_name': client_name,
        'room_name': room.name,
        'room_type': room.room_type,
        'sheets_booked': sheets_booked,
        'price_per_sheet': price_per_sheet,
        'total_price': total_price,
        'discount': discount,
        'discount_amount': discount_amount,
        'final_cost': final_cost,
        'igst': igst,
        'sgst': sgst,
        'total_price_with_gst': total_price_with_gst,
        'start_date': start_date,
        'end_date': end_date,
        'current_date': now().strftime('%Y-%m-%d'),
    }

    # Render the HTML invoice
    html_invoice = render_to_string('invoice_template.html', context)

    # Generate PDF
    pdf = HTML(string=html_invoice).write_pdf()

    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="invoice_{id}.pdf"'

    return response


def paginate_queryset(request, queryset, per_page=10):
    paginator = Paginator(queryset, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return page_obj

def edit_booking(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id)
    
    if request.method == 'POST':
        # Manually handle form submission
        client_name = request.POST.get('client_name')
        room = booking.room  # You can choose to change the room if needed
        sheets_booked = int(request.POST.get('sheets_booked'))
        
        # Update the booking details
        booking.client_name = client_name
        booking.sheets_booked = sheets_booked
        # Optionally, calculate the new final price based on updated details
        booking.final_price = sheets_booked * booking.room.price_per_sheet - booking.discount
        booking.save()

        # Redirect to the booking details or another page after saving
        return redirect('employee:dashboard')  # Or wherever you want to redirect after editing
    
    return render(request, 'edit_booking.html', {'booking': booking})




def cancel_booking(request, booking_id):
    booking = get_object_or_404(Booking, id=booking_id)
    booking.status = 'CANCELED'
    booking.save()
    messages.success(request, f"Booking for {booking.client_name} has been canceled.")
    return redirect('employee:dashboard')



def clients_report(request):
    status = request.GET.get('status')  # ACTIVE, CANCELED, COMPLETED
    export = request.GET.get('export')  # Check if export is requested

    # Filter bookings based on the status
    bookings = Booking.objects.all()
    if status:
        bookings = bookings.filter(status=status)

    # If export to Excel is requested
    if export == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=clients_report.xlsx'

        # Create a new Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clients Report"

        # Add headers
        headers = [
            "Client Name", "Room Name", "Booking Type",
            "Sheets Booked", "Status", "Start Date", "End Date"
        ]
        ws.append(headers)

        # Add booking data
        for booking in bookings:
            ws.append([
                booking.client_name,
                booking.room.name,
                booking.booking_type,
                booking.sheets_booked,
                booking.status,
                booking.start_date,
                booking.end_date
            ])

        # Save the workbook to the response
        wb.save(response)
        return response

    # Handle normal report display
    page_obj = paginate_queryset(request, bookings, per_page=10)
    return render(request, 'reports/clients_report.html', {'bookings': bookings, 'page_obj': page_obj})

def quotation_report(request):
    quotation = Quotation.objects.all()
    # Handle normal report display
    page_obj = paginate_queryset(request, quotation, per_page=10)
    return render(request, 'reports/quotation_report.html', {'quotation': quotation, 'page_obj': page_obj})

def rooms_report(request):
    # Get all rooms
    rooms = Room.objects.all()
    
    # Prepare a list of rooms with booked and available sheets data
    rooms_with_data = []

    for room in rooms:
        # Get all bookings for the room
        bookings = Booking.objects.filter(room=room)
        
        # Calculate booked sheets by summing up the 'sheets_booked' for all bookings
        booked_sheets = bookings.aggregate(Sum('sheets_booked'))['sheets_booked__sum'] or 0
        
        # Calculate available sheets
        available_sheets = room.sheet_count - booked_sheets
        
        # Append the data along with the room
        rooms_with_data.append({
            'room': room,
            'booked_sheets': booked_sheets,
            'available_sheets': available_sheets
        })
    
    # Paginate the rooms with data
    paginator = Paginator(rooms_with_data, 10)  # 10 rooms per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'reports/rooms_report.html', {'page_obj': page_obj})


def available_rooms_report(request):
    rooms = Room.objects.all()
    available_rooms = []

    for room in rooms:
        bookings = Booking.objects.filter(room=room)
        booked_sheets = bookings.aggregate(Sum('sheets_booked'))['sheets_booked__sum'] or 0
        available_sheets = room.sheet_count - booked_sheets
        if available_sheets > 0:
            available_rooms.append({
                'room': room,  # Include the full Room object
                'available_sheets': available_sheets,
                'booked_sheets': booked_sheets,  # Optional: include booked_sheets
            })

    page_obj = paginate_queryset(request, available_rooms, per_page=10)

    return render(request, 'reports/available_rooms_report.html', {'page_obj': page_obj})

def enquiry_report(request):
    
    export = request.GET.get('export')  # Check if export is requested

    # Fetch enquiries with status "Pending"
    enquiries = Enquiry.objects.all()
    # If export to Excel is requested
    if export == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=enquiry_report.xlsx'

        # Create a new Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Enquiry Report"

        # Add headers
        headers = ["Enquiry ID", "Client Name", "Employee Name", "Room Type", "Sheets Requested", 
                   "From Date", "To Date", "Referenced By", "Status", "Remarks", "Created At", "Offer Price"]
        ws.append(headers)

        # Add enquiry data
        for enquiry in enquiries:
            # Convert datetime fields to naive (timezone-less) datetime using make_naive, if applicable
            created_at = enquiry.created_at
            from_date = enquiry.from_date
            to_date = enquiry.to_date

            # Ensure that the datetime objects are naive
            if isinstance(created_at, datetime) and created_at.tzinfo is not None:
                created_at = make_naive(created_at)
            if isinstance(from_date, datetime) and from_date.tzinfo is not None:
                from_date = make_naive(from_date)
            if isinstance(to_date, datetime) and to_date.tzinfo is not None:
                to_date = make_naive(to_date)

            ws.append([
                enquiry.id,
                enquiry.client_name,
                enquiry.employee_name,
                enquiry.room_type,
                enquiry.required_sheets,
                from_date,
                to_date,
                enquiry.referenced_by,
                enquiry.status,
                enquiry.remarks,
                created_at,
                enquiry.offer_price,
            ])

        # Save the workbook to the response
        wb.save(response)
        return response

    # Handle normal enquiry display with pagination
    page_obj = paginate_queryset(request, enquiries, per_page=10)
    context = {
        'enquiries': page_obj.object_list,
        'page_obj': page_obj
    }
    return render(request, 'reports/enquiry_report.html', context)

    
    
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from datetime import datetime, timedelta
from .models import Room, Quotation

# View for listing all quotations
def quotation(request):
    quotations = Quotation.objects.all()
    rooms = Room.objects.all()

    context = {
       
        'quotations': quotations,
        'rooms':rooms
    }
    
    return render(request, 'quotation.html',context)

# View for creating a new quotation

def create_quotation(request):
    if request.method == 'POST':
        room_id = request.POST['room_id']
        client_name = request.POST['client_name']
        sheets_requested = int(request.POST['sheets_requested'])
        booking_type = request.POST['booking_type']
        start_date = datetime.strptime(request.POST['start_date'], '%Y-%m-%d').date()

        # Determine the end date based on booking type
        if booking_type == 'DAILY':
            end_date = datetime.strptime(request.POST['end_date'], '%Y-%m-%d').date()
        elif booking_type == 'MONTHLY':
            end_date = start_date + timedelta(days=30)
        elif booking_type == 'YEARLY':
            end_date = start_date + timedelta(days=365)

        # Fetch the room and check availability
        room = get_object_or_404(Room, id=room_id)
        # Calculate available sheets based on active bookings
        active_bookings = Booking.objects.filter(room=room, status='ACTIVE')
        booked_sheets = active_bookings.aggregate(Sum('sheets_booked'))['sheets_booked__sum'] or 0
        available_sheets = room.sheet_count - booked_sheets


        # If room type is not "Office", don't calculate based on sheets, just store the price
        if room.room_type != "Office":
            total_price = room.price  # Direct price without calculation based on sheets
        else:
            # For "Office" type, calculate price based on sheets requested
            
            if sheets_requested > available_sheets:
                messages.error(request, "Not enough sheets available for the selected dates.")
                return render(request, 'create_quotation.html', {
                    'rooms': Room.objects.all(),
                    'selected_room': room,
                    'room_type': room.room_type,
                    'price': room.price or room.price_per_sheet,
                    'error_sheets_requested': True,  # Pass error flag
                    'available_sheets': available_sheets,  # Ensure this reflects the correct value
                    'request_data': request.POST, 
                })

            price_per_sheet = room.price_per_sheet or 5000
            total_price = sheets_requested * price_per_sheet

        # Discount and GST calculation
        discount = int(request.POST.get('discount', '0') or 0)
        discount_amount = (total_price * discount) / 100
        final_price = total_price - discount_amount

        # Calculate GST
        igst = (final_price * 9) / 100
        sgst = (final_price * 9) / 100
        total_price_with_gst = final_price + igst + sgst

        # Create the booking
        Quotation.objects.create(
            room=room,
            client_name=client_name,
            sheets_booked=sheets_requested if room.room_type == "Office" else 0,  # Set sheets only for "Office"
            start_date=start_date,
            end_date=end_date,
            booking_type=booking_type,
            discount=discount,
            final_price=total_price_with_gst,  # Include GST in the final price
        )

        messages.success(request, f"Room booked successfully from {start_date} to {end_date}!")
        return redirect('employee:dashboard')

    else:
        rooms = Room.objects.all()
        selected_room_id = request.GET.get('room_id')
        selected_room = None
        room_type = None
        price = None
        available_sheets = 0

        if selected_room_id:
            selected_room = Room.objects.get(id=selected_room_id)
            room_type = selected_room.room_type
            price = selected_room.price or selected_room.price_per_sheet

            # Calculate available sheets dynamically
            active_bookings = Booking.objects.filter(room=selected_room, status='ACTIVE')
            booked_sheets = active_bookings.aggregate(Sum('sheets_booked'))['sheets_booked__sum'] or 0
            available_sheets = selected_room.sheet_count - booked_sheets

        else:
            available_sheets = 0

        return render(request, 'create_quotation.html', {
                'rooms': rooms,
                'selected_room': selected_room,
                'room_type': room_type,
                'price': price,
                'error_sheets_requested': False,
                'available_sheets': available_sheets,
            })



   
# View for printing the quotation
def print_quotation(request, quotation_id):
    quotation = get_object_or_404(Quotation, id=quotation_id)
    current_date = date.today()
    return render(request, 'invoice_template.html', {
        'quotation': quotation,
        'current_date': current_date
    })

# View for updating a quotation
def update_quotation(request, id):
    quotation = get_object_or_404(Quotation, id=id)
    
    if request.method == 'POST':
        room = get_object_or_404(Room, id=request.POST['room_id'])
        client_name = request.POST['client_name']
        sheets_requested = int(request.POST['sheets_requested'])
        start_date = request.POST['start_date']
        end_date = request.POST['end_date']
        discount = float(request.POST['discount'])
        
        # Price calculation logic
        price_per_sheet = room.price_per_sheet if room.room_type != 'Auditorium' else room.price
        total_price = sheets_requested * price_per_sheet
        discount_amount = (total_price * discount) / 100
        final_price = total_price - discount_amount

        # GST Calculation
        igst = (final_price * 9) / 100
        sgst = (final_price * 9) / 100
        final_price_with_gst = final_price + igst + sgst
        
        # Update quotation
        quotation.room = room
        quotation.client_name = client_name
        quotation.sheets_requested = sheets_requested
        quotation.start_date = start_date
        quotation.end_date = end_date
        quotation.discount = discount
        quotation.igst = igst
        quotation.sgst = sgst
        quotation.final_price = final_price
        quotation.final_price_with_gst = final_price_with_gst
        quotation.save()

        messages.success(request, "Quotation updated successfully.")
        return redirect('employee:quotation')

    rooms = Room.objects.all()
    return render(request, 'update_quotation.html', {'quotation': quotation, 'rooms': rooms})

# View for deleting a quotation
def delete_quotation(request, id):
    quotation = get_object_or_404(Quotation, id=id)
    quotation.delete()
    messages.success(request, "Quotation deleted successfully.")
    return redirect('employee:quotation')



import openpyxl
from django.http import HttpResponse
from django.utils.timezone import make_naive
from datetime import datetime

def enquiry(request):
    export = request.GET.get('export')  # Check if export is requested

    # Fetch enquiries with status "Pending"
    enquiries = Enquiry.objects.filter(status="Pending")

    # If export to Excel is requested
    if export == 'excel':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=enquiry_report.xlsx'

        # Create a new Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Enquiry Report"

        # Add headers
        headers = ["Enquiry ID", "Client Name", "Employee Name", "Room Type", "Sheets Requested", 
                   "From Date", "To Date", "Referenced By", "Status", "Remarks", "Created At", "Offer Price"]
        ws.append(headers)

        # Add enquiry data
        for enquiry in enquiries:
            # Convert datetime fields to naive (timezone-less) datetime using make_naive, if applicable
            created_at = enquiry.created_at
            from_date = enquiry.from_date
            to_date = enquiry.to_date

            # Ensure that the datetime objects are naive
            if isinstance(created_at, datetime) and created_at.tzinfo is not None:
                created_at = make_naive(created_at)
            if isinstance(from_date, datetime) and from_date.tzinfo is not None:
                from_date = make_naive(from_date)
            if isinstance(to_date, datetime) and to_date.tzinfo is not None:
                to_date = make_naive(to_date)

            ws.append([
                enquiry.id,
                enquiry.client_name,
                enquiry.employee_name,
                enquiry.room_type,
                enquiry.required_sheets,
                from_date,
                to_date,
                enquiry.referenced_by,
                enquiry.status,
                enquiry.remarks,
                created_at,
                enquiry.offer_price,
            ])

        # Save the workbook to the response
        wb.save(response)
        return response

    # Handle normal enquiry display with pagination
    page_obj = paginate_queryset(request, enquiries, per_page=10)
    context = {
        'enquiries': page_obj.object_list,
        'page_obj': page_obj
    }
    return render(request, 'enquiry.html', context)




def create_enquiry(request):
    
    if request.method == 'POST':
        client_name = request.POST.get('client_name')
        employee_name = request.POST.get('employee_name')
        room_type = request.POST.get('room_type')
        required_sheets = request.POST.get('required_sheets')
        from_date = request.POST.get('from_date')
        to_date = request.POST.get('to_date')
        referenced_by = request.POST.get('referenced_by')
        # status = request.POST.get('status', 'Pending')
        remarks = request.POST.get('remarks', '')
        offer_price = request.POST.get('offer_price')

        try:
            # Validate dates
            if from_date > to_date:
                raise ValidationError("From date must be before or equal to To date.")

            # Create the enquiry
            Enquiry.objects.create(
                client_name=client_name,
                employee_name=employee_name,
                room_type=room_type,
                required_sheets=required_sheets,
                from_date=from_date,
                to_date=to_date,
                referenced_by=referenced_by,
                # status=status,
                remarks=remarks,
                offer_price=offer_price,
            )
            messages.success(request, "Enquiry created successfully!")
            return redirect('employee:enquiry')  # Replace with your desired redirect URL
        except Exception as e:
            messages.error(request, f"Error creating enquiry: {e}")

    return render(request, 'create_enquiry.html')
    
def update_enquiry(request, id):
    enquiry = get_object_or_404(Enquiry, id=id)

    if request.method == 'POST':
        # Update the enquiry with form data
        enquiry.client_name = request.POST.get('client_name')
        enquiry.employee_name = request.POST.get('employee_name')
        enquiry.room_type = request.POST.get('room_type')
        enquiry.required_sheets = request.POST.get('required_sheets')
        enquiry.from_date = request.POST.get('from_date')
        enquiry.to_date = request.POST.get('to_date')
        enquiry.referenced_by = request.POST.get('referenced_by')
        enquiry.status = request.POST.get('status')
        enquiry.remarks = request.POST.get('remarks')
        
        # Validate and save
        try:
            enquiry.full_clean()  # Validate the data
            enquiry.save()  # Save to the database
            return redirect('employee:enquiry')  # Redirect to the list page
        except ValidationError as e:
            return render(request, 'enquiry.html', {'enquiry': enquiry, 'errors': e.message_dict})

    
    return render(request, 'update_enquiry.html', {'enquiry': enquiry})


def delete_enquiry(request, id):
    if request.method == 'POST':
        enquiry = get_object_or_404(Enquiry, id=id)
        enquiry.delete()
        return JsonResponse({'success': True, 'message': 'Enquiry deleted successfully.'})
    else:
        return JsonResponse({'success': False, 'message': 'Invalid request method.'})
